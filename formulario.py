import ttkbootstrap as ttk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import re
import os
from tkinter.font import Font
from ttkbootstrap.constants import *

name_archivo = 'datos.xlsx'

# Verificar si el archivo ya existe

if os.path.exists(name_archivo):
    wb = load_workbook(name_archivo)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.append(['Name', 'Age', 'Email', 'Phone', 'Address'])

def save_data():
    name = entry_name.get()
    age = entry_age.get()
    email = entry_email.get()
    phone = entry_phone.get()
    address = entry_address.get()
    
    if not name or not age or not email or not phone or not address:
        messagebox.showwarning(title="Warning", message="All fields are required.")
        return
    try:
        age = int(age)
        phone = int(phone)
    except ValueError:
        messagebox.showwarning(title="Warning", message="Age and Phone must be numbers.")
        return
    
    if not re.match(pattern=r"[^@]+@[^@]+\.[^@]+", string=email):
        messagebox.showwarning(title="Warning", message="The email is not valid.")
        return
    
    ws.append([name, age, email, phone, address])
    wb.save(name_archivo)
    messagebox.showinfo(title="Information", message="Data saved successfully.")
    
    entry_name.delete(0, ttk.END)
    entry_age.delete(0, ttk.END)
    entry_email.delete(0, ttk.END)
    entry_phone.delete(0, ttk.END)
    entry_address.delete(0, ttk.END)
    
    
# Crear ventana principal con ttkbootstrap
root = ttk.Window(themename="darkly")
root.title("Registration Form")

label_title = Font(font="Poppins", size=20, weight="bold")
label_text = Font(font="Poppins", size=10)
label_register = ttk.Label(root, text="REGISTER", font=label_title, anchor=CENTER)
label_register.grid(row=0, column=0, columnspan=2, padx=10, pady=10)

label_name = ttk.Label(root, text="Name:", font=label_text)
label_name.grid(row=1, column=0, padx=10, pady=5)
entry_name = ttk.Entry(root, font=label_text)
entry_name.grid(row=1, column=1, padx=10, pady=5)

label_age = ttk.Label(root, text="Age:", font=label_text)
label_age.grid(row=2, column=0, padx=10, pady=5)
entry_age = ttk.Entry(root, font=label_text)
entry_age.grid(row=2, column=1, padx=10, pady=5)

label_email = ttk.Label(root, text="Email:", font=label_text)
label_email.grid(row=3, column=0, padx=10, pady=5)
entry_email = ttk.Entry(root, font=label_text)
entry_email.grid(row=3, column=1, padx=10, pady=5)

label_phone = ttk.Label(root, text="Phone:", font=label_text)
label_phone.grid(row=4, column=0, padx=10, pady=5)
entry_phone = ttk.Entry(root, font=label_text)
entry_phone.grid(row=4, column=1, padx=10, pady=5)

label_address = ttk.Label(root, text="Address:", font=label_text)
label_address.grid(row=5, column=0, padx=10, pady=5)
entry_address = ttk.Entry(root, font=label_text)
entry_address.grid(row=5, column=1, padx=10, pady=5)

button_save = ttk.Button(root, text="Save", command=save_data)
button_save.grid(row=6, column=0, columnspan=2, padx=10, pady=10)
root.mainloop()